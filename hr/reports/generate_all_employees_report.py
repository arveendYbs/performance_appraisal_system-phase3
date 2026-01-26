#!/usr/bin/env python3
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_all_employees_report(data, output_path):
    """Generate Excel report for all employees"""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Employees Report"
    
    company_name = data.get('company_name', 'Company')
    year = data.get('year', '')
    employees = data.get('employees', [])
    
    # Title styling
    title_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:AZ1')
    ws['A1'] = f'PERFORMANCE ASSESSMENT - ALL EMPLOYEES REPORT'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:AZ2')
    ws['A2'] = f'{company_name} - {year}'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    current_row = 4
    
    # Headers
    base_headers = ['Company', 'Dept', 'Name', 'Staff No.', 'Form', 'Role', 'Position', 
                    'Date Joined', 'Period']
    
    # Find max questions (assume 12)
    max_questions = 12
    
    # Build headers
    headers = base_headers.copy()
    
    # Employee scores
    for i in range(1, max_questions + 1):
        headers.append(f'Q{i}')
    headers.extend(['Total', 'Score', 'Rating'])
    
    # Manager scores
    for i in range(1, max_questions + 1):
        headers.append(f'Q{i}')
    headers.extend(['Total', 'Score', 'Final Rating'])
    
    # Training needs (assume max 10)
    max_training = 10
    for i in range(1, max_training + 1):
        headers.append(f'T{i}')
    
    # Section headers (row before column headers)
    section_row = current_row
    col_idx = len(base_headers) + 1
    
    # Employee scores section
    start_col = col_idx
    end_col = col_idx + max_questions + 2
    ws.merge_cells(start_row=section_row, start_column=start_col, 
                   end_row=section_row, end_column=end_col)
    cell = ws.cell(row=section_row, column=start_col)
    cell.value = 'Performance Assessment - Employee Scores'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    col_idx = end_col + 1
    
    # Manager scores section
    start_col = col_idx
    end_col = col_idx + max_questions + 2
    ws.merge_cells(start_row=section_row, start_column=start_col, 
                   end_row=section_row, end_column=end_col)
    cell = ws.cell(row=section_row, column=start_col)
    cell.value = 'Performance Assessment - Manager Scores'
    cell.font = header_font
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    col_idx = end_col + 1
    
    # Training needs section
    start_col = col_idx
    end_col = col_idx + max_training - 1
    ws.merge_cells(start_row=section_row, start_column=start_col, 
                   end_row=section_row, end_column=end_col)
    cell = ws.cell(row=section_row, column=start_col)
    cell.value = 'Training & Development Needs'
    cell.font = header_font
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
    
    current_row += 1
    
    # Column headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row += 1
    
    # Data rows
    for employee in employees:
        for appraisal in employee.get('appraisals', []):
            col_idx = 1
            
            # Base info
            ws.cell(row=current_row, column=col_idx, value=company_name)
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=employee.get('department', '-'))
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=employee.get('name', '-'))
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=employee.get('emp_number', '-'))
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=appraisal.get('form_title', '-'))
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value='Employee')
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=employee.get('position', '-'))
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=employee.get('date_joined', '-'))
            col_idx += 1
            period = f"{appraisal.get('period_from', '')} to {appraisal.get('period_to', '')}"
            ws.cell(row=current_row, column=col_idx, value=period)
            col_idx += 1
            
            # Employee scores
            questions = appraisal.get('questions', [])
            emp_total = 0
            emp_count = 0
            for i in range(max_questions):
                if i < len(questions):
                    rating = questions[i].get('employee_rating', 0)
                    ws.cell(row=current_row, column=col_idx, value=rating if rating else '')
                    if rating:
                        try:
                            emp_total += float(rating)
                            emp_count += 1
                        except:
                            pass
                else:
                    ws.cell(row=current_row, column=col_idx, value='')
                col_idx += 1
            
            emp_score = round((emp_total / emp_count) * 10, 2) if emp_count > 0 else 0
            emp_grade = get_grade(emp_score)
            
            ws.cell(row=current_row, column=col_idx, value=emp_total)
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=emp_score)
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=emp_grade)
            col_idx += 1
            
            # Manager scores
            mgr_total = 0
            mgr_count = 0
            for i in range(max_questions):
                if i < len(questions):
                    rating = questions[i].get('manager_rating', 0)
                    ws.cell(row=current_row, column=col_idx, value=rating if rating else '')
                    if rating:
                        try:
                            mgr_total += float(rating)
                            mgr_count += 1
                        except:
                            pass
                else:
                    ws.cell(row=current_row, column=col_idx, value='')
                col_idx += 1
            
            mgr_score = round((mgr_total / mgr_count) * 10, 2) if mgr_count > 0 else 0
            
            ws.cell(row=current_row, column=col_idx, value=mgr_total)
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=mgr_score)
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=appraisal.get('grade', '-'))
            col_idx += 1
            
            # Training needs
            training = appraisal.get('training_needs', [])
            for i in range(max_training):
                if i < len(training):
                    ws.cell(row=current_row, column=col_idx, value=training[i])
                else:
                    ws.cell(row=current_row, column=col_idx, value='')
                col_idx += 1
            
            current_row += 1
    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'J6'
    
    # Apply borders to all data
    for row in ws.iter_rows(min_row=5, max_row=current_row-1, 
                           min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    wb.save(output_path)

def get_grade(score):
    """Calculate grade from score"""
    if score >= 85:
        return 'A'
    elif score >= 75:
        return 'B+'
    elif score >= 65:
        return 'B'
    elif score >= 60:
        return 'B-'
    else:
        return 'C'

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python generate_all_employees_report.py <input_json> <output_xlsx>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        create_all_employees_report(data, output_file)
        print(f"Report generated successfully: {output_file}")
        sys.exit(0)
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)