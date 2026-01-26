#!/usr/bin/env python3


import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_report(data_file, output_file):
    try:
        # Load data
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        employee = data['employee']
        appraisals = data['appraisals']
        year = data['year']
        
        print(f"Generating report for: {employee['name']}")
        print(f"Appraisals: {len(appraisals)}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{year} Report"
        
        # Styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        subheader_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        subheader_font = Font(bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Title
        ws['A1'] = f"Summary of {year} Appraisal Ratings - {employee.get('company_name', 'N/A')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:R1')
        current_row = 3
        
        # Section headers
        ws.merge_cells(f'J{current_row}:U{current_row}')
        ws[f'J{current_row}'] = 'Performance Assessment - Employee Scores'
        ws[f'J{current_row}'].fill = subheader_fill
        ws[f'J{current_row}'].font = subheader_font
        ws[f'J{current_row}'].alignment = center_align
        
        ws.merge_cells(f'Y{current_row}:AJ{current_row}')
        ws[f'Y{current_row}'] = 'Performance Assessment - Manager Scores'
        ws[f'Y{current_row}'].fill = subheader_fill
        ws[f'Y{current_row}'].font = subheader_font
        ws[f'Y{current_row}'].alignment = center_align
        
        # Training section header
        ws.merge_cells(f'AN{current_row}:AW{current_row}')
        ws[f'AN{current_row}'] = 'Training & Development Needs'
        ws[f'AN{current_row}'].fill = subheader_fill
        ws[f'AN{current_row}'].font = subheader_font
        ws[f'AN{current_row}'].alignment = center_align
        
        current_row += 1
        
        # Column headers
        headers = [
            'Company', 'Dept', 'Name', 'Staff No.', 'Form', 'Role', 
            'Position', 'Date Joined', 'Period'
        ]
        
        # Employee questions (Q1-Q12)
        for i in range(1, 13):
            headers.append(f'Q{i}')
        
        headers.extend(['Total', 'Score', 'Rating'])
        
        # Manager questions (Q1-Q12)  
        for i in range(1, 13):
            headers.append(f'Q{i}')
        
        headers.extend(['Total', 'Score', 'Final Rating'])
        
        # Training columns (T1-T10 for up to 10 training items)
        for i in range(1, 11):
            headers.append(f'T{i}')
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        
        # Set column widths
        widths = {
            'A': 15, 'B': 12, 'C': 18, 'D': 12, 'E': 20, 'F': 10,
            'G': 18, 'H': 12, 'I': 20
        }
        # Q1-Q12 for employee (J-U)
        for col in range(ord('J'), ord('V')):
            widths[chr(col)] = 6
        # Totals
        widths['V'] = 10
        widths['W'] = 10
        widths['X'] = 10
        # Q1-Q12 for manager (Y-AJ)
        for col_num in range(25, 37):
            widths[get_column_letter(col_num)] = 6
        # Totals
        widths['AK'] = 10
        widths['AL'] = 10
        widths['AM'] = 12
        # Training columns T1-T10 (AN-AW)
        for col_num in range(40, 50):  # AN=40, AW=49
            widths[get_column_letter(col_num)] = 30  # Wide for training names
        
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Write data rows
        for appraisal in appraisals:
            questions = appraisal.get('questions', [])
            
            print(f"Appraisal {appraisal.get('id')}: {len(questions)} questions")
            
            row_data = [
                employee.get('company_name', ''),
                employee.get('department', ''),
                employee.get('name', ''),
                employee.get('emp_number', ''),
                appraisal.get('form_title', ''),
                employee.get('role', '').title(),
                employee.get('position', ''),
                employee.get('date_joined', ''),
                f"{appraisal.get('period_from', '')} to {appraisal.get('period_to', '')}"
            ]
            
            # Employee individual question scores (12 questions)
            for i in range(12):
                if i < len(questions):
                    score = questions[i].get('employee_rating', 0)
                    # Only show non-zero scores
                    row_data.append(score if score else '')
                    print(f"  Emp Q{i+1}: {score}")
                else:
                    row_data.append('')
            
            # Employee total (formula: sum of Q1-Q12)
            row_data.append(f'=SUM(J{current_row}:U{current_row})')
            
            # Employee calculated score (based on role)
            role = employee.get('role', 'employee').lower()
            if 'manager' in role or 'admin' in role:
                divisor = '1.2'
            elif 'worker' in role:
                divisor = '0.8'
            else:
                divisor = '1'
            row_data.append(f'=ROUND(V{current_row}/{divisor},0)')
            
            # Employee rating (A/B+/B/B-/C)
            rating_formula = f'=IF(W{current_row}=0,"",IF(W{current_row}<50,"C",IF(W{current_row}<60,"B-",IF(W{current_row}<75,"B",IF(W{current_row}<85,"B+","A")))))'
            row_data.append(rating_formula)
            
            # Manager individual question scores (12 questions)
            for i in range(12):
                if i < len(questions):
                    score = questions[i].get('manager_rating', 0)
                    # Only show non-zero scores
                    row_data.append(score if score else '')
                    print(f"  Mgr Q{i+1}: {score}")
                else:
                    row_data.append('')
            
            # Manager total (formula: sum of Q1-Q12)
            row_data.append(f'=SUM(Y{current_row}:AJ{current_row})')
            
            # Manager calculated score
            row_data.append(f'=ROUND(AK{current_row}/{divisor},0)')
            
            # Final rating
            final_rating_formula = f'=IF(AL{current_row}=0,"",IF(AL{current_row}<50,"C",IF(AL{current_row}<60,"B-",IF(AL{current_row}<75,"B",IF(AL{current_row}<85,"B+","A")))))'
            row_data.append(final_rating_formula)
            
            # Training items - one per column (T1-T10)
            training_needs = appraisal.get('training_needs', [])
            print(f"  Training needs: {len(training_needs)} items")
            
            for i in range(10):  # 10 training columns
                if i < len(training_needs):
                    row_data.append(training_needs[i])
                    print(f"  T{i+1}: {training_needs[i]}")
                else:
                    row_data.append('')  # Empty if no training item
            
            # Write row
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                # Set alignment
                if col_idx <= 9:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
            
            current_row += 1
        
        # Freeze panes
        ws.freeze_panes = 'A5'
        
        # Save
        wb.save(output_file)
        print(f"SUCCESS: Excel report saved")
        
        import os
        if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
            return True
        return False
            
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python generate-excel-report.py <data_json> <output_xlsx>")
        sys.exit(1)
    
    success = create_report(sys.argv[1], sys.argv[2])
    sys.exit(0 if success else 1)